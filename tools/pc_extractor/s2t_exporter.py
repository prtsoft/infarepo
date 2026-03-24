"""
Source-to-target (S2T) document exporters for MappingLineage objects.

Produces CSV and Excel files in the standard S2T spreadsheet format used
for business sign-off and HIPAA traceability documentation.

HIPAA note: field names and table names are written to the output.
Actual data values are never included — only structural metadata.
"""
from __future__ import annotations

import csv
from pathlib import Path
from typing import TYPE_CHECKING

if TYPE_CHECKING:
    from .models import MappingLineage

# Column headers for the S2T spreadsheet
_HEADERS = [
    "target_table",
    "target_field",
    "source_table",
    "source_field",
    "source_field_type",
    "expression",
    "has_lookup",
    "lookup_names",
    "transformation_chain",
    "notes",
]


def _rows(lineage: "MappingLineage"):
    """Generate one row dict per source reference (or one row for unconnected fields)."""
    for fl in lineage.fields:
        common = {
            "target_table":        fl.target_table,
            "target_field":        fl.target_field,
            "expression":          fl.expression or "",
            "has_lookup":          str(len(fl.lookups) > 0),
            "lookup_names":        "|".join(lkp.lookup_name for lkp in fl.lookups),
            "transformation_chain": " → ".join(
                f"{n.instance}.{n.field}" for n in fl.chain
            ),
            "notes":               "; ".join(fl.notes),
        }
        if fl.sources:
            for src in fl.sources:
                yield {
                    **common,
                    "source_table":      src.table,
                    "source_field":      src.field,
                    "source_field_type": src.field_type,
                }
        else:
            yield {
                **common,
                "source_table":      "",
                "source_field":      "",
                "source_field_type": "",
                "expression":        fl.expression or "(unconnected)",
            }


def write_s2t_csv(lineage: "MappingLineage", output_path: Path) -> None:
    """
    Write lineage as an S2T CSV file.

    Each row represents one (target_field, source_field) pair.
    Unconnected target fields get one row with empty source columns.

    Parameters
    ----------
    lineage:
        MappingLineage result from trace_mapping().
    output_path:
        Destination .csv file path.  Parent directory must exist.
    """
    with open(output_path, "w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=_HEADERS)
        writer.writeheader()
        for row in _rows(lineage):
            writer.writerow(row)


def write_s2t_excel(lineage: "MappingLineage", output_path: Path) -> None:
    """
    Write lineage as a formatted Excel (.xlsx) S2T workbook.

    Requires ``openpyxl``.  The sheet has:
    - Bold header row with column auto-width
    - Frozen top row
    - Alternating row fill for readability
    - Sheet named after the mapping

    Parameters
    ----------
    lineage:
        MappingLineage result from trace_mapping().
    output_path:
        Destination .xlsx file path.  Parent directory must exist.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise ImportError(
            "openpyxl is required for Excel export.  "
            "Install it with: pip install openpyxl"
        ) from exc

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = lineage.mapping_name[:31]  # Excel sheet name limit

    # ── Header row ─────────────────────────────────────────────────────────
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2E4057")
    header_alignment = Alignment(horizontal="center", wrap_text=True)

    for col_idx, header in enumerate(_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    ws.freeze_panes = "A2"

    # ── Data rows ──────────────────────────────────────────────────────────
    alt_fill = PatternFill("solid", fgColor="F0F4F8")

    for row_idx, row in enumerate(_rows(lineage), start=2):
        fill = alt_fill if row_idx % 2 == 0 else None
        for col_idx, header in enumerate(_HEADERS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row.get(header, ""))
            if fill:
                cell.fill = fill

    # ── Column widths ──────────────────────────────────────────────────────
    _COL_WIDTHS = {
        "target_table":          20,
        "target_field":          25,
        "source_table":          20,
        "source_field":          25,
        "source_field_type":     18,
        "expression":            35,
        "has_lookup":             10,
        "lookup_names":          20,
        "transformation_chain":  45,
        "notes":                 30,
    }
    for col_idx, header in enumerate(_HEADERS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = _COL_WIDTHS.get(header, 18)

    wb.save(output_path)
